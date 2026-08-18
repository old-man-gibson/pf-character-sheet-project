"""Dump a worksheet as a coordinate grid so the layout can be inspected.

Usage: python tools/dump_tab.py <workbook.xlsx> <tab name> [--formulas]
"""
import sys
import openpyxl

path = sys.argv[1]
tab = sys.argv[2]
show_formulas = "--formulas" in sys.argv

wb = openpyxl.load_workbook(path, data_only=not show_formulas)
ws = wb[tab]

merged = {}
for rng in ws.merged_cells.ranges:
    merged[(rng.min_row, rng.min_col)] = f"{rng.max_row - rng.min_row + 1}x{rng.max_col - rng.min_col + 1}"

for row in ws.iter_rows():
    cells = []
    for c in row:
        if c.value is None:
            continue
        v = str(c.value).replace("\n", "\\n")
        if len(v) > 90:
            v = v[:87] + "..."
        tag = merged.get((c.row, c.column))
        cells.append(f"{c.coordinate}={v}" + (f" [M{tag}]" if tag else ""))
    if cells:
        print(" | ".join(cells))
