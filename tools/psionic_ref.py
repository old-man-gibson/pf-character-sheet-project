"""Build the shared power-point table from a workbook's psionicRef tab.

    python tools/psionic_ref.py <workbook.xlsx> [-o data/psionics.json]

The tab is a reference table, not character data, and it is the same in every
workbook, so it is extracted once into a file every character shares.

Almost none of it is load-bearing. The tab looks like thirteen class columns, but
the sheet's own formula never reads them:

    HLOOKUP(PPoints1, PowerPoints, ML + 1, false)

`PowerPoints` is `B1:F21` -- five power-point curves, one per column -- and
`PPoints1` is the block's own **PP@20** cell, whose dropdown is `B1:F1`, the row
of level-20 totals. So a manifesting class is chosen by *which curve it uses*,
named by the total it reaches at 20, and the thirteen class columns to the right
are a crib for a human picking from that dropdown. That is also why a homebrew
manifesting class needs no special handling: it was never a class lookup.

Layout:

    row 1       B..F   the five curves' level-20 totals -- their keys
                H..T   thirteen class names
    rows 2-21   B..F   power points at character levels 1-20
                H..T   each class's column, a copy of whichever curve it uses
    A23:A32            Talent, 1st..9th -- the power-level dropdown's options
"""
import argparse
import json
import os
import re
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

CURVE_COLUMNS = range(2, 7)      # B..F
CLASS_COLUMNS = range(8, 21)     # H..T
LEVELS = 20

# The sheet writes a dash for a level a class cannot manifest at yet, which is
# not the same as manifesting with zero points in the pool.
PLACEHOLDER = re.compile(r"^[�–—-]+$")


def text(v):
    return "" if v is None else str(v).strip()


def number(v):
    s = text(v)
    if s == "" or PLACEHOLDER.match(s):
        return None
    try:
        f = float(s)
    except ValueError:
        return None
    return int(f) if f.is_integer() else f


def read_tables(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    if "psionicRef" not in wb.sheetnames:
        sys.exit(f"{path} has no psionicRef tab")
    ws = wb["psionicRef"]
    at = lambda r, c: ws.cell(r, c).value      # noqa: E731 -- a name for one lookup

    curves = []
    for col in CURVE_COLUMNS:
        total = number(at(1, col))
        if total is None:
            continue
        points = [number(at(1 + lvl, col)) for lvl in range(1, LEVELS + 1)]
        curves.append({"total": total, "points": points})

    # Each class column is a copy of one of the curves; its level-20 cell is the
    # total that identifies which, and so the value its PP@20 dropdown wants.
    classes = []
    for col in CLASS_COLUMNS:
        name = text(at(1, col))
        if not name:
            continue
        total = number(at(1 + LEVELS, col))
        if total is None:
            continue
        classes.append({"name": name, "total": total})

    # The power-level dropdown's own options, in the tab's order, read from the
    # column that carries them rather than assumed.
    power_levels = []
    for row in range(1, ws.max_row + 1):
        v = text(at(row, 1))
        if v == "Talent":
            while True:
                v = text(at(row, 1))
                if not v:
                    break
                power_levels.append(v)
                row += 1
            break

    wb.close()
    return curves, classes, power_levels


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", help="any workbook built on the campaign template")
    parser.add_argument("-o", "--out", default="data/psionics.json")
    args = parser.parse_args()

    curves, classes, power_levels = read_tables(args.workbook)
    if not curves:
        sys.exit("no power-point curves found in B1:F21")

    os.makedirs(os.path.dirname(args.out) or ".", exist_ok=True)
    with open(args.out, "w", encoding="utf-8") as f:
        json.dump({"powerLevels": power_levels, "curves": curves, "classes": classes},
                  f, separators=(",", ":"), ensure_ascii=False)

    size = os.path.getsize(args.out)
    print(f"{len(curves)} curves, {len(classes)} classes -> {args.out} ({size:,} b)")
    for c in curves:
        pts = c["points"]
        print(f"  total {c['total']:>4}  level 1: {pts[0]}  level 20: {pts[-1]}")
    print(f"  power levels: {', '.join(power_levels)}")
    by_total = {}
    for c in classes:
        by_total.setdefault(c["total"], []).append(c["name"])
    for total, names in sorted(by_total.items()):
        print(f"  {total:>4}: {', '.join(names)}")


if __name__ == "__main__":
    main()
