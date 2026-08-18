"""Convert exported Pathfinder character workbooks into the app's JSON schema.

The workbooks all descend from one template ("Bear's sheet"), so extraction is
driven by two robust mechanisms rather than hard-coded cell addresses:

  1. Defined names. The template declares ~476 named ranges (StrMod, BAB,
     CondEntangled, ...). These are stable across characters and give us every
     scalar the sheet author considered meaningful.
  2. Label-anchored table scans. Tables (skills, gear, planner rows) are located
     by finding their header label, then walking rows beneath it.

Google-only formulas (ARRAYFORMULA/FILTER) survive the xlsx export as
`__xludf.DUMMYFUNCTION` placeholders, so formulas are not recoverable for every
cell -- but the *cached values* are. We therefore read values, and reimplement
the derived-stat maths natively in the app's formula engine.

Usage: python tools/convert.py path/to/workbook.xlsx --id name
       python tools/convert.py --out private/characters --raw private/raw   (rebuild a roster)
"""

import argparse
import json
import os
import re
import sys
import datetime
import openpyxl

# Where workbooks are read from and documents written to by default. The
# published app bundles no characters (each visitor adds their own, kept in
# their browser), so a bundled roster is whatever a deployment chooses to put
# in OUT_DIR/index.json -- there is no list of characters in this file.
RAW_DIR = "data/raw"
OUT_DIR = "data/characters"

ABILITIES = ["Str", "Dex", "Con", "Int", "Wis", "Cha"]

# Reference/lookup tabs: machinery, not character data.
REF_TABS = {
    "InstructionsProviso", "vancianRef", "dataSheet", "maneuversRef",
    "psionicRef", "techRef",
}

# Reference tabs that are nonetheless captured cell-for-cell: techRef is a
# character's own technique catalogue (one column per technique, an approval
# status on each), which the model reads into its techniques block on load
# (importTechniques in app/js/model.js). Its named ranges are still skipped.
CAPTURED_REF_TABS = {"techRef"}

# Tabs handled by a dedicated extractor; everything else visible is captured
# generically so nothing silently disappears.
#
# The Template tab is not one of them: it is captured cell-for-cell like any
# other tab and read into feature groups by the model (`importTemplateTab` in
# app/js/model.js), which keeps that scan in one place rather than in both
# converters.
STRUCTURED_TABS = {
    "Character Info", "Stats", "Planner", "Feats", "Mythic", "Equipment",
    "Background & Lore", "Combat Training", "Magic Training",
}


def slug(s):
    """A filename- and URL-safe id, matching the app's own slug rule."""
    return re.sub(r"^_+|_+$", "", re.sub(r"[^a-z0-9]+", "_", str(s or "").lower())) or "x"


def clean(v):
    """Normalise a cell value: trim strings, collapse float-ints, drop blanks."""
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        if v == "" or v.startswith("#"):  # #REF!, #NAME?, ...
            return None
        return v
    if isinstance(v, float) and v.is_integer():
        return int(v)
    if isinstance(v, datetime.datetime):
        return v.isoformat()
    return v


def num(v, default=0):
    v = clean(v)
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, str):
        m = re.match(r"^[+\s]*(-?\d+(?:\.\d+)?)", v.replace(",", ""))
        if m:
            f = float(m.group(1))
            return int(f) if f.is_integer() else f
    return default


class Book:
    """Thin wrapper giving label- and name-based access to a workbook."""

    def __init__(self, path):
        self.wb = openpyxl.load_workbook(path, data_only=True)
        self.grids = {}
        for ws in self.wb.worksheets:
            self.grids[ws.title] = [[clean(c.value) for c in row] for row in ws.iter_rows()]

    def tabs(self):
        return [(ws.title, ws.sheet_state) for ws in self.wb.worksheets]

    def has(self, tab):
        return tab in self.grids

    def cell(self, tab, row, col):
        """1-indexed row/col."""
        g = self.grids.get(tab)
        if not g or row > len(g) or row < 1:
            return None
        r = g[row - 1]
        if col > len(r) or col < 1:
            return None
        return r[col - 1]

    def row(self, tab, row):
        g = self.grids.get(tab)
        if not g or row > len(g) or row < 1:
            return []
        return g[row - 1]

    def find_label(self, tab, label, max_row=None):
        """Return (row, col) of the first cell exactly matching `label`."""
        g = self.grids.get(tab)
        if not g:
            return None
        for i, r in enumerate(g):
            if max_row and i >= max_row:
                break
            for j, v in enumerate(r):
                if isinstance(v, str) and v == label:
                    return (i + 1, j + 1)
        return None

    def right_of(self, tab, label, offset=1):
        """Value `offset` columns right of a label cell -- the sheet's dominant idiom."""
        pos = self.find_label(tab, label)
        if not pos:
            return None
        return self.cell(tab, pos[0], pos[1] + offset)

    def named(self):
        """Resolve every defined name to a scalar or small 2D block."""
        out = {}
        for name in self.wb.defined_names:
            dn = self.wb.defined_names[name]
            val = dn.value
            if not isinstance(val, str) or val.startswith("LAMBDA"):
                continue
            try:
                dests = list(dn.destinations)
            except Exception:
                continue
            for tab, ref in dests:
                if tab not in self.grids or tab in REF_TABS:
                    continue
                m = re.match(r"^\$?([A-Z]+)\$?(\d+)(?::\$?([A-Z]+)\$?(\d+))?$", ref)
                if not m:
                    continue
                c1, r1, c2, r2 = m.groups()
                col1 = openpyxl.utils.column_index_from_string(c1)
                row1 = int(r1)
                if c2 is None:
                    v = self.cell(tab, row1, col1)
                    if v is not None:
                        out[name] = v
                else:
                    col2 = openpyxl.utils.column_index_from_string(c2)
                    row2 = int(r2)
                    # Skip giant lookup blocks; they are rules data, not character data.
                    if (row2 - row1 + 1) * (col2 - col1 + 1) > 400:
                        continue
                    block = [
                        [self.cell(tab, r, c) for c in range(col1, col2 + 1)]
                        for r in range(row1, row2 + 1)
                    ]
                    if any(any(x is not None for x in r) for r in block):
                        out[name] = block
                break
        return out


def strip_row(row):
    """Trim trailing Nones so rows compare cleanly."""
    r = list(row)
    while r and r[-1] is None:
        r.pop()
    return r


def table(bk, tab, header_row, start_row, col_start, col_end, stop_blank=3):
    """Read a rectangular table into a list of dicts keyed by the header row."""
    if not bk.has(tab):
        return []
    headers = []
    for c in range(col_start, col_end + 1):
        h = bk.cell(tab, header_row, c)
        headers.append(str(h) if h is not None else f"col{c}")
    rows = []
    blanks = 0
    r = start_row
    g = bk.grids[tab]
    while r <= len(g):
        vals = [bk.cell(tab, r, c) for c in range(col_start, col_end + 1)]
        if all(v is None for v in vals):
            blanks += 1
            if blanks >= stop_blank:
                break
            r += 1
            continue
        blanks = 0
        rows.append({h: v for h, v in zip(headers, vals) if v is not None})
        r += 1
    return rows


# --------------------------------------------------------------------------
# Section extractors
# --------------------------------------------------------------------------

def extract_identity(bk):
    CI = "Character Info"
    speeds = []
    for r in range(4, 8):
        t = bk.cell(CI, r, 13)  # M: movement type
        if t is None:
            continue
        speeds.append({
            "type": str(t),
            "base": num(bk.cell(CI, r, 14)),
            "bonus": num(bk.cell(CI, r, 15)),
            "final": num(bk.cell(CI, r, 16)),
        })
    return {
        "name": bk.cell(CI, 3, 3),
        "player": bk.cell(CI, 3, 9),
        "race": bk.cell(CI, 4, 3),
        "size": bk.cell(CI, 4, 6),
        "gender": bk.cell(CI, 4, 9),
        "age": bk.cell(CI, 4, 11),
        "variant": bk.cell(CI, 5, 3),
        "heroPoints": {"current": num(bk.cell(CI, 5, 5)), "max": num(bk.cell(CI, 5, 7), 3)},
        "height": bk.cell(CI, 5, 9),
        "weight": bk.cell(CI, 5, 11),
        "level": num(bk.cell(CI, 6, 3)),
        "alignment": bk.cell(CI, 6, 6),
        "deity": bk.cell(CI, 6, 9),
        "specialty": bk.cell(CI, 7, 3),
        "specialtyFeat": bk.cell(CI, 7, 6) or bk.cell(CI, 7, 8),
        "specialtyPerks": [x for x in [bk.cell(CI, 8, 3), bk.cell(CI, 8, 8)] if x],
        "image": bk.cell(CI, 4, 18),
        "nativeLanguages": bk.cell(CI, 10, 3),
        "downtimeLanguages": bk.cell(CI, 10, 8),
        "intLanguages": bk.cell(CI, 11, 3),
        "linguisticsLanguages": bk.cell(CI, 12, 3),
        "proficiencies": {
            "weapons": bk.cell(CI, 11, 15),
            "armor": bk.cell(CI, 12, 15),
            "shield": bk.cell(CI, 13, 15),
        },
        "speeds": speeds,
        "mythicPath": bk.cell(CI, 18, 14),
        "mythicTier": num(bk.cell(CI, 19, 14)),
        "focusStat": bk.cell(CI, 20, 14),
        "guild": bk.cell(CI, 15, 17),
        "primordiaTechnique": bk.cell(CI, 16, 14),
    }


def extract_abilities(bk):
    """Ability scores plus the Stats-tab breakdown of where each bonus came from."""
    CI, ST = "Character Info", "Stats"
    breakdown_headers = []
    if bk.has(ST):
        breakdown_headers = [bk.cell(ST, 2, c) for c in range(3, 20)]

    out = {}
    for i, ab in enumerate(ABILITIES):
        r = 15 + i
        entry = {
            "score": num(bk.cell(CI, r, 3)),
            "mod": num(bk.cell(CI, r, 4)),
            "tempScore": num(bk.cell(CI, r, 5)),
            "totalMod": num(bk.cell(CI, r, 6)),
            "checkMod": num(bk.cell(CI, r, 7)),
        }
        if bk.has(ST):
            src = {}
            for j, h in enumerate(breakdown_headers):
                if not h:
                    continue
                v = bk.cell(ST, 3 + i, 3 + j)
                if v is not None and v is not False and v != 0:
                    src[str(h)] = v
            entry["sources"] = src
        out[ab.lower()] = entry
    return out


# Stats-tab column letters -> build keys. Column B holds the computed total.
STATS_COLUMNS = {
    3: "pointBuy",     # C
    4: "race",         # D
    5: "abp",          # E  Automatic Bonus Progression enhancement
    6: "gear",         # F  gear enhancement (caps with abp at +6)
    7: "attunement",   # G  boolean in the sheet; +2 when set
    8: "inherent",     # H
    9: "array",        # I  optional array
    10: "level4",      # J  the every-fourth-level increase
    11: "mythic",      # K
    12: "size",        # L  permanent
    13: "untyped",     # M  permanent
    15: "alchemical",  # O  temporary from here down
    16: "circumstance",  # P
    17: "morale",      # Q
    18: "tempEnhancement",  # R
    19: "tempSize",    # S
}


def extract_stats_build(bk):
    """
    The Stats tab: every bonus that feeds an ability score, by source.

    Reproduces as:
      enhancement = min(6, abp + gear)
      total       = pointBuy + race + enhancement + attunement + inherent
                    + array + level4 + mythic + size + untyped
      tempTotal   = total + alchemical + circumstance + morale
                    + tempEnhancement + tempSize
    """
    ST = "Stats"
    if not bk.has(ST):
        return None
    build = {}
    for i, ab in enumerate(ABILITIES):
        r = 3 + i
        entry = {}
        for col, key in STATS_COLUMNS.items():
            v = bk.cell(ST, r, col)
            if key == "attunement":
                entry[key] = 2 if v is True else num(v)
            else:
                entry[key] = num(v)
        entry["sheetTotal"] = num(bk.cell(ST, r, 2))
        build[ab.lower()] = entry
    return build


def extract_point_buy_table(bk):
    """dataSheet K21:L33 -- ability score to point-buy cost."""
    if not bk.has("dataSheet"):
        return None
    table_ = {}
    for r in range(22, 40):
        score = bk.cell("dataSheet", r, 11)   # K
        cost = bk.cell("dataSheet", r, 12)    # L
        if isinstance(score, (int, float)) and isinstance(cost, (int, float)):
            table_[int(score)] = int(cost)
    return table_ or None


def _planner_column(bk, prefix):
    """Find a Planner column by the start of its header text.

    The Prowess/Array block sits at a different offset in different characters'
    sheets (AN/AO/AP in some, AL/AM/AN in others), so it must be located by
    label rather than by a fixed column index.
    """
    for c in range(1, 80):
        h = bk.cell("Planner", 1, c)
        if isinstance(h, str) and h.strip().startswith(prefix):
            return c
    return None


def extract_progression_picks(bk):
    """
    Ability-boosting choices recorded on the Planner, by level.

      "Level/4"           the every-fourth-level +1 increase
      Mental Prowess      ABP, +2 each
      Physical Prowess    ABP, +2 each
      "Array (Optional)"  the optional array, +2 each, 4 slots per row

    Planner row = level + 1.
    """
    if not bk.has("Planner"):
        return None

    col_l4 = _planner_column(bk, "Level/4") or 2
    col_mental = _planner_column(bk, "Mental Prowess")
    col_physical = _planner_column(bk, "Physical Prowess")
    col_array = _planner_column(bk, "Array")

    level4, abp, array = [], [], []
    for level in range(1, 21):
        r = level + 1
        pick = bk.cell("Planner", r, col_l4)
        if isinstance(pick, str) and pick.strip():
            level4.append({"level": level, "ability": pick.strip()})

        mental = bk.cell("Planner", r, col_mental) if col_mental else None
        physical = bk.cell("Planner", r, col_physical) if col_physical else None
        if isinstance(mental, str) or isinstance(physical, str):
            abp.append({
                "level": level,
                "mental": mental.strip() if isinstance(mental, str) else None,
                "physical": physical.strip() if isinstance(physical, str) else None,
            })

        if col_array:
            slots = [bk.cell("Planner", r, c) for c in range(col_array, col_array + 4)]
            slots = [s.strip() if isinstance(s, str) else None for s in slots]
            if any(slots):
                array.append({"level": level, "slots": slots})

    return {
        "level4": level4,
        "abp": abp,
        "array": array,
        "arrayNote": (bk.cell("Planner", 1, col_array) if col_array else None),
    }


def extract_classes(bk):
    CI = "Character Info"
    out = []
    for r in range(23, 41):
        name = bk.cell(CI, r, 2)
        if not name or (isinstance(name, str) and re.match(r"^Class \d+$", name)):
            continue
        out.append({
            "name": name,
            "hd": num(bk.cell(CI, r, 4)),
            "bab": bk.cell(CI, r, 5),
            "babOverride": bk.cell(CI, r, 6),
            "goodFort": bool(bk.cell(CI, r, 7)),
            "goodRef": bool(bk.cell(CI, r, 8)),
            "goodWill": bool(bk.cell(CI, r, 9)),
            "skillRanks": num(bk.cell(CI, r, 10)),
            "archetypes": bk.cell(CI, r, 11),
        })
    return out


def extract_conditions(bk):
    """The two condition columns (K/L and M/N) on Character Info."""
    CI = "Character Info"
    out = {}
    for r in range(44, 53):
        for lbl_c, val_c in ((11, 12), (13, 14)):
            lbl = bk.cell(CI, r, lbl_c)
            if lbl:
                out[str(lbl)] = num(bk.cell(CI, r, val_c))
    # The shared template has Helpless and Paralyzed at 1 in every workbook,
    # with nothing else on: a leftover, not a state. Cleared so a fresh import
    # does not open with the character paralysed.
    on = sorted(k for k, v in out.items() if v)
    if on == ["Helpless", "Paralyzed"]:
        for k in on:
            out[k] = 0
    return out


def extract_defenses(bk):
    CI = "Character Info"
    return {
        "ac": num(bk.cell(CI, 44, 3)),
        "touch": num(bk.cell(CI, 45, 3)),
        "flatFooted": num(bk.cell(CI, 46, 3)),
        "cmd": num(bk.cell(CI, 47, 3)),
        "ffCmd": num(bk.cell(CI, 47, 5)),
        "acStat1": bk.cell(CI, 44, 5),
        "acStat2": bk.cell(CI, 45, 5),
        "bonusAC": bk.cell(CI, 46, 5),
        "uncannyDodge": bool(bk.cell(CI, 44, 9)),
        "nonTouch": num(bk.cell(CI, 45, 8)),
        "miscAC": num(bk.cell(CI, 46, 8)),
        "miscCMD": num(bk.cell(CI, 47, 8)),
        "spellResistance": bk.cell(CI, 48, 3),
        "dr": bk.cell(CI, 48, 5),
        "weakness": bk.cell(CI, 48, 8),
        "immunities": bk.cell(CI, 49, 5),
        "resistance": bk.cell(CI, 50, 5),
    }


def extract_attack(bk):
    CI = "Character Info"
    rows = [
        ("melee", 53), ("altMelee", 54), ("ranged", 55),
        ("altRanged", 56), ("cmb", 57), ("altCmb", 58),
    ]
    modes = {}
    for key, r in rows:
        modes[key] = {
            "value": num(bk.cell(CI, r, 3)),
            "stat1": bk.cell(CI, r, 5),
            "stat2": bk.cell(CI, r, 7),
        }
    return {
        "bab": num(bk.cell(CI, 52, 3)),
        "iterative": bk.cell(CI, 52, 5),
        "miscBonus": num(bk.cell(CI, 52, 9)),
        "modes": modes,
        "totalMelee": num(bk.cell(CI, 54, 9)),
        "totalRanged": num(bk.cell(CI, 56, 9)),
        "totalCmb": num(bk.cell(CI, 58, 9)),
    }


def extract_saves(bk):
    CI = "Character Info"
    out = {}
    for key, r in (("fortitude", 61), ("reflex", 62), ("will", 63)):
        out[key] = {
            "total": num(bk.cell(CI, r, 3)),
            "stat1": bk.cell(CI, r, 5),
            "stat2": bk.cell(CI, r, 7),
            "base": num(bk.cell(CI, r, 10)),
        }
    return out


def extract_traits(bk):
    CI = "Character Info"
    traits = []
    for r in range(60, 74):
        label = bk.cell(CI, r, 11)
        if not label:
            continue
        kind = bk.cell(CI, r, 12)
        desc = bk.cell(CI, r, 14)
        if kind is None and desc is None:
            continue
        traits.append({"slot": str(label), "category": kind, "text": desc})
    race_traits = []
    for r in range(66, 74):
        v = bk.cell(CI, r, 3)
        if v:
            race_traits.append(v)
    return traits, race_traits


def extract_skills(bk):
    """Skills table: label-anchored so a shifted template still lines up."""
    CI = "Character Info"
    pos = bk.find_label(CI, "Skills")
    if not pos:
        return []
    header_row = pos[0]
    sub = header_row + 1  # Level / Specialty / Gear / Other / ... sub-headers
    rank_headers = [bk.cell(CI, sub, c) for c in range(7, 15)]

    out = []
    r = header_row + 2
    g = bk.grids[CI]
    blanks = 0
    while r <= len(g):
        name = bk.cell(CI, r, 2)
        if name is None:
            blanks += 1
            if blanks >= 3:
                break
            r += 1
            continue
        blanks = 0
        if str(name) in ("Skills",):
            r += 1
            continue
        # The rank-budget metrics live directly under the skill rows; they are
        # not skills, and everything from there on belongs to other sections.
        if str(name).startswith(("Bonus Skill points", "Int Bonus per",
                                 "Total Skill Points", "Other Trackables",
                                 "Name, Info")):
            break
        ranks = {}
        for j, h in enumerate(rank_headers):
            if not h:
                continue
            v = num(bk.cell(CI, r, 7 + j), None)
            if v:
                ranks[str(h)] = v
        abil = [bk.cell(CI, r, c) for c in (17, 18, 19)]
        out.append({
            "name": str(name),
            "spec": bk.cell(CI, r, 3),
            "bonus": num(bk.cell(CI, r, 4)),
            "classSkill": bool(bk.cell(CI, r, 5)),
            "totalRanks": num(bk.cell(CI, r, 6)),
            "ranks": ranks,
            "requiresTraining": str(bk.cell(CI, r, 15) or "").lower() == "yes",
            "armorPenalty": bool(bk.cell(CI, r, 16)),
            "abilities": [a for a in abil if a],
            "situational": bk.cell(CI, r, 20),
        })
        r += 1
    return out


def extract_skill_budget(bk):
    """The rank-budget metrics under the skills table: bonus skill points per
    level and the Int bonus per level, used to validate assigned ranks."""
    CI = "Character Info"
    pos = bk.find_label(CI, "Bonus Skill points per Level")
    if not pos:
        return {"bonusPerLevel": 0, "intPerLevel": 0}
    r, c = pos
    out = {
        "bonusPerLevel": num(bk.cell(CI, r, c + 1)),
        "intPerLevel": 0,
        "sheetPerLevel": 0,
    }
    g = bk.grids[CI]
    for rr in range(r, min(r + 4, len(g) + 1)):
        lbl = bk.cell(CI, rr, c)
        if isinstance(lbl, str) and lbl.startswith("Int Bonus per"):
            out["intPerLevel"] = num(bk.cell(CI, rr, c + 1))
        if isinstance(lbl, str) and lbl.startswith("Total Skill Points per"):
            out["sheetPerLevel"] = num(bk.cell(CI, rr, c + 1))
    return out


def extract_carry(bk):
    CI = "Character Info"
    out = {"tiers": []}
    for r in range(44, 49):
        lbl = bk.cell(CI, r, 22)
        if lbl:
            out["tiers"].append({"name": str(lbl), "limit": num(bk.cell(CI, r, 24))})
    out["antHaul"] = num(bk.cell(CI, 49, 23), 1)
    out["strBonus"] = num(bk.cell(CI, 50, 23))
    out["quadruped"] = bool(bk.cell(CI, 51, 23))
    out["carried"] = num(bk.cell(CI, 52, 23))
    return out


def extract_resources(bk):
    """The sheet's own Resource Tracker -- seeds of the custom tracker system."""
    CI = "Character Info"
    pos = bk.find_label(CI, "Resource Tracker")
    if not pos:
        return []
    hdr = pos[0] + 1
    c = pos[1]
    out = []
    r = hdr + 1
    while r <= hdr + 20:
        name = bk.cell(CI, r, c)
        if name:
            out.append({
                "name": str(name),
                "uses": num(bk.cell(CI, r, c + 1)),
                "total": bk.cell(CI, r, c + 2),
                "refresh": bk.cell(CI, r, c + 3),
            })
        r += 1
    return out


def extract_hp(bk):
    CI = "Character Info"
    return {
        "total": num(bk.cell(CI, 18, 10)),
        "ability": bk.cell(CI, 18, 12),
        "fcb": num(bk.cell(CI, 19, 10)),
        "ability2": bk.cell(CI, 19, 12),
        "toughness": num(bk.cell(CI, 20, 10)),
        "initiative": num(bk.cell(CI, 15, 9)),
        "initAbility": bk.cell(CI, 15, 10),
        "initAbility2": bk.cell(CI, 15, 11),
    }


def extract_planner(bk):
    if not bk.has("Planner"):
        return []
    headers = {}
    for c in range(1, 60):
        h = bk.cell("Planner", 1, c)
        if h:
            headers[c] = str(h)
    out = []
    for r in range(2, 22):
        row = {}
        for c, h in headers.items():
            v = bk.cell("Planner", r, c)
            if v is not None:
                row[h] = v
        if row:
            out.append(row)
    return out


def extract_feats(bk):
    """Feats tab is a set of side-by-side category columns."""
    if not bk.has("Feats"):
        return {}
    groups = {}
    for c in range(1, 40):
        cat = bk.cell("Feats", 2, c)
        if not cat:
            continue
        entries = []
        for r in range(3, 25):
            name = bk.cell("Feats", r, c)
            if not name or str(name) in ("Name",):
                continue
            meta = bk.cell("Feats", r, c + 2)
            entries.append({"name": str(name), "detail": meta})
        if entries:
            key = str(cat)
            n = 2
            base = key
            while key in groups:
                key = f"{base} {n}"
                n += 1
            groups[key] = entries
    return groups


def extract_mythic(bk):
    if not bk.has("Mythic"):
        return {}
    out = {
        "path": bk.right_of("Mythic", "Mythic Path"),
        "tier": num(bk.right_of("Mythic", "Mythic Tier")),
        "baseAbilities": [],
        "abilities": [],
        "tradition": {},
        "flowingPower": False,
    }
    for r in range(6, 10):
        v = bk.cell("Mythic", r, 3)
        if v and str(v) not in ("Base Path Ability", "Base Mythic Abilities"):
            out["baseAbilities"].append(v)
    out["basePathAbility"] = bk.right_of("Mythic", "Base Path Ability")

    # Two layouts: the current template has a Lvl column (labels in C), the
    # older one starts the ability slots in column B.
    new_layout = bk.cell("Mythic", 11, 2) == "Lvl"
    c_slot = 3 if new_layout else 2

    r_trad = None
    for r in range(12, 40):
        if bk.cell("Mythic", r, c_slot) == "Mythic Tradition":
            r_trad = r
            break

    for r in range(12, (r_trad or 32)):
        slot = bk.cell("Mythic", r, c_slot)
        name = bk.cell("Mythic", r, c_slot + 1)
        feat = bk.cell("Mythic", r, c_slot + 3)
        extra = bk.cell("Mythic", r, c_slot + 4)
        stat = bk.cell("Mythic", r, c_slot + 5)
        if not (name or feat or extra or stat):
            continue
        out["abilities"].append({
            "level": num(bk.cell("Mythic", r, 2), None) if new_layout else None,
            "slot": slot,
            "name": name,
            "path": bk.cell("Mythic", r, c_slot + 2),
            "feat": feat,
            "featChoice": extra,
            "statBonus": stat,
        })

    # Mythic tradition: one mandatory drawback unlocking a boon, up to two more
    # drawbacks for two more boons, and a quality.
    if r_trad:
        keys = {"Drawback 1": "drawback1", "Drawback 2": "drawback2",
                "Drawback 3": "drawback3", "Quality": "quality",
                "Boon 1": "boon1", "Boon 2": "boon2", "Boon 3": "boon3"}
        for r in range(r_trad + 1, r_trad + 9):
            lbl = bk.cell("Mythic", r, c_slot)
            if isinstance(lbl, str) and lbl in keys:
                out["tradition"][keys[lbl]] = bk.cell("Mythic", r, c_slot + 1)
        # Flowing Power checkbox sits beside the header in the new layout.
        fp = bk.cell("Mythic", r_trad, c_slot + 3)
        out["flowingPower"] = bool(fp) if isinstance(fp, bool) else False
    return out


GEAR_SLOTS = ["Head", "Headband", "Eyes", "Shoulders", "Neck", "Chest", "Body",
              "Armor", "Belt", "Wrists", "Hands", "Ring 1", "Ring 2", "Feet"]


def _gear_row(bk, tab, r, slot):
    """One slotted-gear row: name, three typed bonuses, four other bonuses."""
    bonuses = []
    for (vc, tc) in ((3, 4), (5, 6), (7, 8)):
        v = bk.cell(tab, r, vc)
        t = bk.cell(tab, r, tc)
        bonuses.append({"value": num(v) if v is not None else None,
                        "type": t if isinstance(t, str) else None})
    others = [bk.cell(tab, r, c) for c in (9, 11, 13, 15)]
    return {
        "slot": slot,
        "name": bk.cell(tab, r, 2),
        "bonuses": bonuses,
        "others": [o if isinstance(o, str) else (str(o) if o is not None else None) for o in others],
        "weight": num(bk.cell(tab, r, 17)),
        "cost": num(bk.cell(tab, r, 18)),
    }


def _armor_row(bk, tab, r, kind):
    return {
        "kind": kind,
        "name": bk.cell(tab, r, 2),
        "acBonus": num(bk.cell(tab, r, 3)),
        "maxDex": num(bk.cell(tab, r, 4), None),
        "acp": num(bk.cell(tab, r, 5)),
        "type": bk.cell(tab, r, 6),
        "ghostTouch": bool(bk.cell(tab, r, 7)),
        "spellFailure": bk.cell(tab, r, 8),
        "others": [x for x in (bk.cell(tab, r, 9), bk.cell(tab, r, 11),
                               bk.cell(tab, r, 13), bk.cell(tab, r, 15)) if x],
        "weight": num(bk.cell(tab, r, 17)),
        "cost": num(bk.cell(tab, r, 18)),
    }


def extract_equipment(bk):
    """Equipment worksheet: slotted gear, other items, armor & shields, and up
    to six weapon blocks. Every section is label-anchored — some characters
    have extra shield rows that shift everything below them."""
    tab = "Equipment"
    if not bk.has(tab):
        return {"gear": [], "other": [], "armor": None, "shields": [], "weapons": []}

    gear = [_gear_row(bk, tab, 2 + i, s) for i, s in enumerate(GEAR_SLOTS)]
    other = [_gear_row(bk, tab, 17 + i, f"Other {i + 1}") for i in range(8)]

    # Armor block: header row "Armor | Armor Value | Max Dex ...", then an
    # Armor row and one or more Shield rows.
    armor = None
    shields = []
    r_hdr = None
    for r in range(20, 36):
        if bk.cell(tab, r, 1) == "Armor" and bk.cell(tab, r, 3) == "Armor Value":
            r_hdr = r
            break
    if r_hdr:
        armor = _armor_row(bk, tab, r_hdr + 1, "Armor")
        r = r_hdr + 2
        while isinstance(bk.cell(tab, r, 1), str) and bk.cell(tab, r, 1).startswith("Shield"):
            shields.append(_armor_row(bk, tab, r, bk.cell(tab, r, 1)))
            r += 1

    # Weapon blocks, anchored on their "Weapon Name N" header rows.
    weapons = []
    g = bk.grids[tab]
    for r_name in range(1, len(g) + 1):
        hdr = bk.cell(tab, r_name, 1)
        if not (isinstance(hdr, str) and re.match(r"^Weapon Name \d+$", hdr)):
            continue
        row1, row2, row3 = r_name + 1, r_name + 3, r_name + 4
        name = bk.cell(tab, row1, 1)
        if not name and not bk.cell(tab, row1, 5):
            continue
        weapons.append({
            "name": name,
            "attackType": bk.cell(tab, row1, 3),
            "sheetAttack": num(bk.cell(tab, row1, 4), None),
            "dice": bk.cell(tab, row1, 5),
            "damageAbility": bk.cell(tab, row1, 6),
            "abilityMult": num(bk.cell(tab, row1, 7), 1) or 1,
            "miscDamage": num(bk.cell(tab, row1, 8)),
            "sheetTotalDamage": bk.cell(tab, row1, 9),
            "critRange": num(bk.cell(tab, row1, 10), None),
            "critMult": bk.cell(tab, row1, 11),
            "bonusCritDamage": bk.cell(tab, row1, 12),
            "damageType": bk.cell(tab, row1, 13),
            "groups": [bk.cell(tab, row2, 1), bk.cell(tab, row2, 2), bk.cell(tab, row2, 3)],
            "miscAttack": num(bk.cell(tab, row2, 4)),
            "special": bk.cell(tab, row2, 5),
            "ammunition": bk.cell(tab, row2, 9),
            "size": bk.cell(tab, row2, 10),
            "range": bk.cell(tab, row2, 13),
            "enhancement": num(bk.cell(tab, row3, 2)),
            "familiarity": bk.cell(tab, row3, 4),
            "handedness": bk.cell(tab, row3, 6),
            "weight": num(bk.cell(tab, row3, 8)),
            "price": num(bk.cell(tab, row3, 10)),
        })

    return {
        "gear": gear,
        "other": other,
        "armor": armor,
        "shields": shields,
        "weapons": weapons,
    }


def extract_background(bk):
    if not bk.has("Background & Lore"):
        return {}
    out = {}
    g = bk.grids["Background & Lore"]
    for r in range(1, len(g) + 1):
        for c in range(1, 12):
            lbl = bk.cell("Background & Lore", r, c)
            if not isinstance(lbl, str):
                continue
            if lbl in ("Personality", "Appearance", "Likes", "Dislikes", "Goals",
                       "Fears", "Character Strengths", "Character Flaws",
                       "Friends/Family", "Enemies/Rivals", "Additional Information",
                       "Ez'atian Certifications"):
                chunks = []
                for rr in range(r + 1, min(r + 10, len(g) + 1)):
                    v = bk.cell("Background & Lore", rr, c)
                    if v:
                        chunks.append(str(v))
                out[lbl] = "\n".join(chunks) if chunks else None
    return out


# ---------------------------------------------------------------------------
# Spheres training (Combat Training / Magic Training)
#
# Both tabs share one layout: up to three class blocks side by side, each with
# a per-level row (rows 5-24 = levels 1-20). Cached values (cumulative talents,
# granted flags, CL) are extracted alongside the entered talent names so the
# app can verify its own progression maths against the sheet's.
# ---------------------------------------------------------------------------

# (name, type, talents/level, mod1, mod2, countCol, grantedCol, progCol, nameCol, sphereCol)
TRAINING_BLOCKS = [
    ((1, 5), (2, 6), (4, 6), (3, 6), (3, 7), 2, 3, 4, 5, 7),      # E/K block 1
    ((1, 11), (2, 12), (4, 12), (3, 12), (3, 13), 8, 9, 10, 11, 13),  # class 2
    ((1, 17), (2, 18), (4, 18), (3, 18), (3, 19), 14, 15, 16, 17, 19),  # class 3
]


# Blocks 4-6 sit on the extended (level 21-40) page. Their per-level rows are
# out of scope, but their headers still name casting classes, which count for
# the tradition spell-point multiplier.
TRAINING_BLOCKS_EXT = [
    ((26, 5), (27, 6), (29, 6), (28, 6), (28, 7)),
    ((26, 11), (27, 12), (29, 12), (28, 12), (28, 13)),
    ((26, 17), (27, 18), (29, 18), (28, 18), (28, 19)),
]


def _training_classes(bk, tab):
    classes = []
    for (nm, ty, tpl, m1, m2, c_cnt, c_grant, c_prog, c_name, c_sph) in TRAINING_BLOCKS:
        name = bk.cell(tab, *nm)
        if not name:
            continue
        levels = []
        for lvl in range(1, 21):
            r = 4 + lvl
            levels.append({
                "level": lvl,
                "sheetCount": num(bk.cell(tab, r, c_cnt)),
                "sheetGranted": bool(bk.cell(tab, r, c_grant)),
                "sheetProgression": num(bk.cell(tab, r, c_prog)),
                "talent": bk.cell(tab, r, c_name),
                "sphere": bk.cell(tab, r, c_sph),
            })
        classes.append({
            "name": str(name),
            "type": bk.cell(tab, *ty),
            "talentsPerLevel": bk.cell(tab, *tpl),
            "mod1": bk.cell(tab, *m1),
            "mod2": bk.cell(tab, *m2),
            "levels": levels,
        })
    for (nm, ty, tpl, m1, m2) in TRAINING_BLOCKS_EXT:
        name = bk.cell(tab, *nm)
        if not name:
            continue
        classes.append({
            "name": str(name),
            "type": bk.cell(tab, *ty),
            "talentsPerLevel": bk.cell(tab, *tpl),
            "mod1": bk.cell(tab, *m1),
            "mod2": bk.cell(tab, *m2),
            "levels": [],
            "extended": True,
        })
    return classes


def _grid_texts(bk, tab, r1, r2, c1, c2):
    out = []
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            v = bk.cell(tab, r, c)
            if isinstance(v, str) and v.strip():
                out.append(v.strip())
    return out


def _label_row(bk, tab, col, label, start=1, end=80):
    """Row of the first cell in `col` whose text starts with `label`.

    The tradition/global blocks sit on slightly different rows across template
    revisions, so every read is anchored to its label rather than a fixed row.
    """
    for r in range(start, end + 1):
        v = bk.cell(tab, r, col)
        if isinstance(v, str) and v.strip().startswith(label):
            return r
    return None


def extract_combat_training(bk):
    tab = "Combat Training"
    if not bk.has(tab):
        return None
    U, V, W, X = 21, 22, 23, 24

    r_mt = _label_row(bk, tab, U, "Martial Tradition") or 4
    tradition = {"name": bk.cell(tab, r_mt, V), "entries": []}
    r_bt = _label_row(bk, tab, U, "Bonus Talents") or (r_mt + 15)
    for r in range(r_mt + 2, r_bt - 1):
        t = bk.cell(tab, r, U)
        s = bk.cell(tab, r, X)
        if t or s:
            tradition["entries"].append({"talent": t, "sphere": s})

    bonus = []
    r_br = _label_row(bk, tab, U, "Bonus Ranks") or (r_bt + 27)
    for r in range(r_bt + 1, r_br - 1):
        t = bk.cell(tab, r, U)
        if t:
            bonus.append({
                "talent": t,
                "sphere": bk.cell(tab, r, W),
                "source": bk.cell(tab, r, X),
            })

    sphere_bonuses = []
    r_ss = _label_row(bk, tab, 26, "Sphere")
    if r_ss:
        for r in range(r_ss + 1, r_ss + 27):
            s = bk.cell(tab, r, 26)
            if s:
                sphere_bonuses.append({
                    "sphere": str(s),
                    "rankBonus": num(bk.cell(tab, r, 27)),
                    "dcBonus": num(bk.cell(tab, r, 28)),
                    "sheetValue": bk.cell(tab, r, 29),
                })

    skill_ranks = []
    for r in range(r_br + 1, r_br + 18):
        s = bk.cell(tab, r, U)
        if not s:
            continue
        enabled = bk.cell(tab, r, V)
        skill_ranks.append({
            "skill": str(s),
            "enabled": True if enabled is None else bool(num(enabled, 0) or enabled is True),
            "multiplier": num(bk.cell(tab, r, 25), 1) or 1,
            "sheetCurrent": num(bk.cell(tab, r, W)),
            "sheetMax": num(bk.cell(tab, r, X)),
        })

    def vrow(label, col=V, default=0):
        r = _label_row(bk, tab, U, label)
        return num(bk.cell(tab, r, col)) if r else default

    r_dice = _label_row(bk, tab, U, "Unarmed Strike dice")
    r_kn = _label_row(bk, tab, U, "Talented Knuckle")
    r_st = _label_row(bk, tab, U, "Step Increases")
    r_sz = _label_row(bk, tab, U, "Size Increases")
    r_ub = _label_row(bk, tab, U, "Uses Boxing")
    r_uo = _label_row(bk, tab, U, "Uses Open Hand")
    other = []
    for r in [r_kn, r_st, r_sz]:
        if r:
            for c in (W, X):
                v = bk.cell(tab, r, c)
                if isinstance(v, str) and v.strip() and not v.strip().startswith(("Uses", "Other", "Unorth")):
                    other.append(v.strip())

    unarmed = {
        "sheetDice": bk.cell(tab, r_dice, V) if r_dice else None,
        "talentedKnuckle": vrow("Talented Knuckle"),
        "brawlersVest": vrow("Brawler's Vest"),
        "stepIncreases": vrow("Step Increases"),
        "sizeIncreases": vrow("Size Increases"),
        "usesBoxing": bool(num(bk.cell(tab, r_ub, V), 1)) if r_ub else True,
        "usesBrute": bool(num(bk.cell(tab, r_ub, X), 1)) if r_ub else True,
        "usesOpenHand": bool(num(bk.cell(tab, r_uo, V), 1)) if r_uo else True,
        "usesWrestling": bool(num(bk.cell(tab, r_uo, X), 1)) if r_uo else True,
        "otherSpheres": other,
        "veilEssence": num(bk.cell("Akashic", 35, 21)) if bk.has("Akashic") else 0,
    }

    r_dc = _label_row(bk, tab, 28, "DC Base")
    return {
        "classes": _training_classes(bk, tab),
        "tradition": tradition,
        "bonusTalents": bonus,
        "sphereBonuses": sphere_bonuses,
        "skillRanks": skill_ranks,
        "unarmed": unarmed,
        "sheetBaseDC": num(bk.cell(tab, r_dc, 29)) if r_dc else 0,
    }


def extract_magic_training(bk):
    tab = "Magic Training"
    if not bk.has(tab):
        return None
    U, V, W = 21, 22, 23

    # Drawbacks fill every row between the Tradition header and the Boons
    # line — templates differ in how many rows that is. Bought-off drawbacks
    # follow, under a "Drawback Feats" heading in newer revisions.
    r_trad = _label_row(bk, tab, U, "Tradition /") or 2
    r_boons = _label_row(bk, tab, U, "Boons:") or (r_trad + 6)
    drawbacks = _grid_texts(bk, tab, r_trad + 1, r_boons - 1, U, W)
    def rows_until_blank(r1, r2):
        out = []
        for r in range(r1, r2 + 1):
            texts = [bk.cell(tab, r, c) for c in range(U, W + 1)]
            texts = [t.strip() for t in texts if isinstance(t, str) and t.strip()]
            if not texts or any(t.startswith("Casting Bonus") for t in texts):
                break
            out.extend(texts)
        return out

    r_feats = _label_row(bk, tab, U, "Drawback Feats", start=r_boons)
    if r_feats:
        bought_off = rows_until_blank(r_feats + 1, r_feats + 5)
    else:
        bought_off = rows_until_blank(r_boons + 1, r_boons + 5)
    bought_off = [b for b in bought_off if not b.startswith("Drawback Feats")]

    bonus = []
    r_bt = _label_row(bk, tab, U, "Casting Bonus Spheres")
    if r_bt:
        for r in range(r_bt + 1, r_bt + 5):
            t = bk.cell(tab, r, U)
            if t:
                bonus.append({"talent": t, "sphere": bk.cell(tab, r, W)})

    sphere_bonuses = []
    r_ss = _label_row(bk, tab, 25, "Sphere")   # Y column header
    if r_ss:
        for r in range(r_ss + 1, r_ss + 27):
            s = bk.cell(tab, r, 25)
            if s:
                sphere_bonuses.append({
                    "sphere": str(s),
                    "clBonus": num(bk.cell(tab, r, 26)),
                    "dcBonus": num(bk.cell(tab, r, 27)),
                    "sheetValue": bk.cell(tab, r, 28),
                })

    def block(label):
        """(bonusValue, sheetTotal) for a Base X / Bonus X / Total X block."""
        r = _label_row(bk, tab, U, label)
        if not r:
            return 0, 0
        return num(bk.cell(tab, r + 1, V)), num(bk.cell(tab, r + 1, W))

    dc_bonus, dc_total = block("Base Global DC")
    cl_bonus, cl_total = block("Global Caster Level")
    msb_bonus, msb_total = block("Base MSB")
    msd_bonus, msd_total = block("Base MSD")

    def at_label(label, col=V):
        r = _label_row(bk, tab, U, label)
        return num(bk.cell(tab, r, col)) if r else 0

    r_amt = _label_row(bk, tab, U, "AMT")
    r_mamt = _label_row(bk, tab, U, "Mythic AMT")
    r_sp1 = _label_row(bk, tab, U, "Class 1 SP")

    return {
        "classes": _training_classes(bk, tab),
        "sphereBonuses": sphere_bonuses,
        "tradition": {
            "name": bk.cell(tab, r_trad, V),
            "drawbacks": drawbacks,
            "boughtOff": bought_off,
        },
        "bonusTalents": bonus,
        "amt": bool(bk.cell(tab, r_amt, V)) if r_amt else False,
        "mythicAmt": bool(bk.cell(tab, r_mamt, V)) if r_mamt else False,
        "dcBonus": dc_bonus,
        "clBonus": cl_bonus,
        "msbBonus": msb_bonus,
        "msdBonus": msd_bonus,
        "bonusSP": at_label("Bonus SP"),
        "sheet": {
            "boons": num(bk.cell(tab, r_boons, V)) if r_boons else 0,
            "spTier": num(bk.cell(tab, r_boons, W)) if r_boons else 0,
            "totalDC": dc_total,
            "totalCL": cl_total,
            "totalMSB": msb_total,
            "totalMSD": msd_total,
            "traditionSP": at_label("Tradition SP"),
            "totalSP": at_label("Total SP"),
            "classSP": [num(bk.cell(tab, r_sp1 + i, V)) for i in range(6)] if r_sp1 else [],
        },
    }


def extract_generic(bk, tab):
    """Fallback: keep a tab's non-empty cells so nothing is lost in conversion."""
    g = bk.grids.get(tab)
    if not g:
        return None
    rows = []
    for i, r in enumerate(g):
        rr = strip_row(r)
        if any(v is not None for v in rr):
            rows.append({"r": i + 1, "cells": rr})
    return {"rows": rows}


def convert(path, key, title, file_id):
    bk = Book(path)
    traits, race_traits = extract_traits(bk)

    tabs = bk.tabs()
    extra = {}
    for name, state in tabs:
        if (name in REF_TABS and name not in CAPTURED_REF_TABS) or name in STRUCTURED_TABS:
            continue
        if name in ("Character Info",):
            continue
        gen = extract_generic(bk, name)
        if gen and gen["rows"]:
            extra[name] = {"hidden": state != "visible", **gen}

    doc = {
        # Bump whenever the shape changes: the app discards saved local edits
        # whose schemaVersion does not match, rather than silently loading a
        # document that is missing newly added sections.
        "schemaVersion": 9,
        "id": key,
        "source": {
            "title": title,
            "fileId": file_id,
            "url": (f"https://docs.google.com/spreadsheets/d/{file_id}/edit"
                    if file_id else ""),
            "convertedAt": datetime.datetime.now().isoformat(timespec="seconds"),
        },
        "tabs": [{"name": n, "hidden": s != "visible"} for n, s in tabs],
        "identity": extract_identity(bk),
        "abilities": extract_abilities(bk),
        "statsBuild": extract_stats_build(bk),
        "pointBuyTable": extract_point_buy_table(bk),
        "progressionPicks": extract_progression_picks(bk),
        "hp": extract_hp(bk),
        "classes": extract_classes(bk),
        "defenses": extract_defenses(bk),
        "conditions": extract_conditions(bk),
        "attack": extract_attack(bk),
        "saves": extract_saves(bk),
        "traits": traits,
        "raceTraits": race_traits,
        "skills": extract_skills(bk),
        "skillBudget": extract_skill_budget(bk),
        "carry": extract_carry(bk),
        "resources": extract_resources(bk),
        "planner": extract_planner(bk),
        "feats": extract_feats(bk),
        "mythic": extract_mythic(bk),
        "equipment": extract_equipment(bk),
        "background": extract_background(bk),
        "training": {
            "combat": extract_combat_training(bk),
            "magic": extract_magic_training(bk),
        },
        "named": bk.named(),
        "extraTabs": extra,
    }
    return doc


def index_entry(key, doc):
    """The row index.json carries for a character: what the picker needs."""
    ident = doc["identity"]
    return {
        "id": key,
        "name": ident["name"],
        "race": ident["race"],
        "level": ident["level"],
        "classes": [c["name"] for c in doc["classes"]],
        "image": ident["image"],
        "file": f"{key}.json",
    }


def summary_line(key, doc, size):
    ident = doc["identity"]
    return (f"{key:10s} {str(ident['name'])[:28]:30s} L{ident['level']:<3} "
            f"skills={len(doc['skills']):3d} named={len(doc['named']):3d} "
            f"extra={len(doc['extraTabs'])} {size:>8,d}b")


def warnings_for(doc):
    """Flag the things a sheet off the template is likely to be missing.

    The extractors return empty structures rather than failing, so a workbook
    that is not the campaign template converts to a thin document instead of an
    error. These lines are what tells you that happened.
    """
    out = []
    ident = doc["identity"]
    if not ident.get("name"):
        out.append("no character name found (Character Info tab missing or renamed?)")
    if not ident.get("level"):
        out.append("no character level found")
    if not doc["skills"]:
        out.append("no skills table found")
    if not doc["classes"]:
        out.append("no classes found")
    if not doc["named"]:
        out.append("no defined names - this does not look like the campaign template")
    if not any(r.get("Level") for r in doc["planner"]):
        out.append("Planner is empty, so the Progression tab will start blank")
    return out


def write_character(key, doc, out_dir):
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, f"{key}.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(doc, f, indent=1, ensure_ascii=False)
    return path, os.path.getsize(path)


def upsert_index(entry, out_dir):
    """Add or replace one row in index.json, keeping the existing order."""
    path = os.path.join(out_dir, "index.json")
    rows = []
    if os.path.exists(path):
        with open(path, encoding="utf-8") as f:
            rows = json.load(f).get("characters", [])
    for i, row in enumerate(rows):
        if row.get("id") == entry["id"]:
            rows[i] = entry
            break
    else:
        rows.append(entry)
    with open(path, "w", encoding="utf-8") as f:
        json.dump({"characters": rows}, f, indent=1, ensure_ascii=False)
    return len(rows)


def convert_all(out_dir, raw_dir):
    """Rebuild every character listed in out_dir/index.json, and the index with them.

    The roster is the existing index, and each character's sheet title and
    Google file id come from the document already on disk, so nothing about
    who the characters are lives in this script.
    """
    index_path = os.path.join(out_dir, "index.json")
    if not os.path.exists(index_path):
        print(f"error: no {index_path} to rebuild from -- convert a workbook by name first")
        return 1
    with open(index_path, encoding="utf-8") as f:
        roster = json.load(f).get("characters", [])
    index = []
    for row in roster:
        key = row["id"]
        raw = os.path.join(raw_dir, f"{key}.xlsx")
        if not os.path.exists(raw):
            print(f"skipped  {key}: no {raw}")
            continue
        prior = {}
        try:
            with open(os.path.join(out_dir, f"{key}.json"), encoding="utf-8") as f:
                prior = json.load(f).get("source", {}) or {}
        except (json.JSONDecodeError, OSError):
            pass
        if not prior.get("fileId"):
            print(f"note     {key}: no recorded Sheets link -- re-run with --file-id to restore it")
        doc = convert(raw, key, prior.get("title") or row.get("name") or key,
                      prior.get("fileId") or "")
        _, size = write_character(key, doc, out_dir)
        index.append(index_entry(key, doc))
        print(summary_line(key, doc, size))

    os.makedirs(out_dir, exist_ok=True)
    with open(os.path.join(out_dir, "index.json"), "w", encoding="utf-8") as f:
        json.dump({"characters": index}, f, indent=1, ensure_ascii=False)
    print(f"\nWrote {len(index)} characters + index to {out_dir}/")
    return 0


def convert_one(args):
    """Convert a single workbook and slot it into index.json."""
    path = args.workbook
    if not os.path.exists(path):
        print(f"error: no such file: {path}")
        return 1
    if not path.lower().endswith((".xlsx", ".xlsm")):
        print(f"error: expected an .xlsx workbook, got {os.path.basename(path)}")
        return 1

    stem = os.path.splitext(os.path.basename(path))[0]
    key = args.id or slug(stem)
    out_dir = args.out

    # Re-converting an updated workbook keeps the title and Sheets link already
    # recorded, so a plain re-run does not quietly downgrade them to defaults.
    prior = {}
    existing = os.path.join(out_dir, f"{key}.json")
    if os.path.exists(existing):
        try:
            with open(existing, encoding="utf-8") as f:
                prior = json.load(f).get("source", {}) or {}
        except (json.JSONDecodeError, OSError):
            pass
    title = args.name or prior.get("title") or stem
    file_id = args.file_id or prior.get("fileId") or ""

    print(f"reading  {path}")
    doc = convert(path, key, title, file_id)
    structured = sum(1 for t in doc["tabs"] if t["name"] in STRUCTURED_TABS)
    print(f"tabs     {len(doc['tabs'])} ({structured} structured, "
          f"{len(doc['extraTabs'])} captured verbatim)")

    for w in warnings_for(doc):
        print(f"warning  {w}")

    if args.dry_run:
        print(f"dry run  would write {os.path.join(out_dir, key + '.json')} "
              f"and index it as \"{key}\"")
        return 0

    verb = "replaced" if os.path.exists(existing) else "wrote"
    written, size = write_character(key, doc, out_dir)
    count = upsert_index(index_entry(key, doc), out_dir)
    print(f"{verb:8s} {written}  ({size:,d}b)")
    print(f"updated  {os.path.join(out_dir, 'index.json')}  ({count} characters)")
    return 0


def main():
    # Character names carry macrons and quotes ("Dokei Saburo" is spelled with
    # two o-macrons), which a legacy Windows console encoding cannot print --
    # and an UnicodeEncodeError in a progress line would abort the conversion.
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, OSError):
        pass

    parser = argparse.ArgumentParser(
        description="Convert Pathfinder character workbooks into the app's JSON.",
        epilog="With no workbook, rebuilds every character listed in the output "
               "directory's index.json from the raw directory and rewrites the index.")
    parser.add_argument("workbook", nargs="?",
                        help="path to an .xlsx workbook; omit to rebuild the whole roster")
    parser.add_argument("--id", help="character id and filename stem "
                                     "(default: the workbook's filename)")
    parser.add_argument("--name", help="source sheet title recorded in the document "
                                       "(default: the workbook's filename)")
    parser.add_argument("--file-id", dest="file_id",
                        help="Google Sheets file id, to record a link back to the source")
    parser.add_argument("--out", default=OUT_DIR,
                        help=f"output directory (default: {OUT_DIR})")
    parser.add_argument("--raw", default=RAW_DIR,
                        help=f"where a roster rebuild finds <id>.xlsx (default: {RAW_DIR})")
    parser.add_argument("--dry-run", action="store_true",
                        help="report what would be written without writing it")
    args = parser.parse_args()

    if not args.workbook:
        for flag in ("id", "name", "file_id"):
            if getattr(args, flag):
                parser.error(f"--{flag.replace('_', '-')} only applies to a single workbook")
        if args.dry_run:
            parser.error("--dry-run only applies to a single workbook")
        return convert_all(args.out, args.raw)
    return convert_one(args)


if __name__ == "__main__":
    raise SystemExit(main())
