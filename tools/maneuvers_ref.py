"""Build the shared discipline catalogue from a workbook's maneuversRef tab.

    python tools/maneuvers_ref.py <workbook.xlsx> [-o data/extensions/path-of-war-disciplines.json]

The tab is a reference table, not character data: it is byte-identical in every
workbook (38 disciplines, 3,873 cells), so it is extracted once into a file every
character shares rather than being copied into each of them. A character records
only which disciplines it knows and which maneuvers are ticked; the names and
types come from here.

Layout, the same shape the Maneuvers tab uses:

    row 1     discipline name, repeated across its three columns
    row 3     "Maneuver" / "Type" headers, one pair per discipline
    row 4+    entries, with the level in column A and the section
              ("Maneuvers" / "Stances") in column B, each carrying down
              until the next one appears
"""
import argparse
import os
import re
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from extension_pack import write_pack  # noqa: E402

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

LEVEL_RE = re.compile(r"^(\d+)(?:st|nd|rd|th) Level$")


def text(v):
    return "" if v is None else str(v).strip()


def read_catalogue(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    if "maneuversRef" not in wb.sheetnames:
        sys.exit(f"{path} has no maneuversRef tab")
    ws = wb["maneuversRef"]
    grid = [[c.value for c in row] for row in ws.iter_rows()]

    def at(r, c):
        return grid[r][c] if 0 <= r < len(grid) and 0 <= c < len(grid[r]) else None

    # The header row is the one carrying the Maneuver/Type pairs; each pair's
    # own column holds the known flag, so a discipline starts one to its left.
    header_row = next(
        (r for r in range(len(grid))
         if any(text(at(r, c)) == "Maneuver" and text(at(r, c + 1)) == "Type"
                for c in range(len(grid[r])))),
        -1,
    )
    if header_row < 0:
        sys.exit("no Maneuver/Type header row found")

    columns = sorted({
        c - 1 for c in range(len(grid[header_row]))
        if text(at(header_row, c)) == "Maneuver" and text(at(header_row, c + 1)) == "Type"
    })

    disciplines = []
    for col in columns:
        name = text(at(header_row - 3, col)) or text(at(header_row - 2, col)) \
            or text(at(header_row - 1, col))
        if not name:
            continue
        entries = []
        level, kind = 0, "maneuver"
        for r in range(header_row + 1, len(grid)):
            lvl = LEVEL_RE.match(text(at(r, 0)))
            if lvl:
                level = int(lvl.group(1))
            section = text(at(r, 1))
            if section.startswith("Maneuver"):
                kind = "maneuver"
            elif section.startswith("Stance"):
                kind = "stance"
            entry = text(at(r, col + 1))
            if not entry:
                continue
            entries.append({
                "level": level,
                "kind": kind,
                "name": entry,
                "type": text(at(r, col + 2)),
            })
        if entries:
            disciplines.append({"name": name, "entries": entries})
    return disciplines


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", help="any workbook built on the campaign template")
    parser.add_argument("-o", "--out", default="data/extensions/path-of-war-disciplines.json")
    args = parser.parse_args()

    disciplines = read_catalogue(args.workbook)
    write_pack(args.out, "path-of-war-disciplines", "maneuvers", {"disciplines": disciplines},
               name="Path of War disciplines",
               description="The discipline catalogue: every maneuver and stance each discipline grants, by level.")

    total = sum(len(d["entries"]) for d in disciplines)
    size = os.path.getsize(args.out)
    print(f"{len(disciplines)} disciplines, {total} maneuvers -> {args.out} ({size:,} b)")
    for d in disciplines[:4]:
        print(f"  {d['name']:22s} {len(d['entries']):3d} entries")
    print("  ...")


if __name__ == "__main__":
    main()
