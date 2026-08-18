"""Build the shared casting table from a workbook's vancianRef tab.

    python tools/vancian_ref.py <workbook.xlsx> [-o data/vancian.json]

The tab is a reference table, not character data: it is the same in every
workbook of a given template revision, so it is extracted once into a file every
character shares rather than being copied into each of them. A character records
only which class each casting block draws its table from; the numbers come from
here.

Layout -- three stacked tables sharing one set of columns:

    row 1     class name, then "<class> 2".."<class> 10" across its ten columns
    row 2     "0th".."9th", the spell-level labels -- the marker a block starts on
    rows 3+   spells per day, one row per class level 1-20
    row 24/25 the same headers again, then spells known for class levels 1-20
    row 47/48 the same headers again, then bonus slots for class levels 1-20

The blocks are **not** in the same order as the class list in column A, which is
only the dropdown's options: column A's second entry is Druid while the second
block is Sorcerer. The sheet's own formula finds a block by HLOOKUP on its row-1
header and so does this, or 32 of the 34 classes would come out holding another
class's table.

Bonus slots exist for only the two classes whose formula asks for them (Cleric's
domain slots and Legendary Medium's), and the sheet writes them by concatenating
a suffix onto the slot count -- "4 +1". They are kept as their own number here.
"""
import argparse
import json
import os
import re
import sys

try:
    import openpyxl
    from openpyxl.utils import range_boundaries
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

# The two named ranges the sheet's own formula went through. Both were left
# behind when classes were appended to the tab, so the last few blocks became
# unreachable on the sheet -- see `reachability` below.
DROPDOWN_NAME = "VancianClasses"
LOOKUP_NAME = "VancianLookup"

# The ten spell-level labels a block spans, and the label its first column
# carries -- which is what marks a block start.
SPELL_LEVEL_LABELS = ["0th", "1st", "2nd", "3rd", "4th", "5th", "6th", "7th", "8th", "9th"]
BLOCK_MARKER = SPELL_LEVEL_LABELS[0]
SPELL_LEVELS = len(SPELL_LEVEL_LABELS)
CLASS_LEVELS = 20

# The three tables, in the order they are stacked. The sheet names only the
# first ("Spells per day" in column A); the rest are identified by position,
# which is what its OFFSET constants (+2, +25, +48 from a block's row 1) encode.
TABLES = ["perDay", "known", "bonus"]

# A trailing " 2".." 10" is the block's own column numbering, not part of the name.
COLUMN_SUFFIX = re.compile(r"\s+(?:[2-9]|10)$")

# The sheet leaves an em dash, or an unreadable glyph where one was, for a spell
# level the class cannot cast at all -- which is not the same as zero slots.
PLACEHOLDER = re.compile(r"^[�–—-]+$")


def text(v):
    return "" if v is None else str(v).strip()


def number(v):
    """A slot count, or None where the class cannot cast at that level."""
    s = text(v)
    if s == "" or PLACEHOLDER.match(s):
        return None
    try:
        f = float(s)
    except ValueError:
        return s          # "+1", "(+3)" -- the bonus table's own notation
    return int(f) if f.is_integer() else f


def boundaries(wb, name):
    """A defined name's cell range as (min_col, min_row, max_col, max_row)."""
    dn = wb.defined_names.get(name)
    if dn is None:
        return None
    ref = str(dn.value).split("!")[-1].replace("$", "")
    try:
        min_col, min_row, max_col, max_row = range_boundaries(ref)
    except ValueError:
        return None
    return min_col, min_row, max_col, max_row


def read_tables(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    if "vancianRef" not in wb.sheetnames:
        sys.exit(f"{path} has no vancianRef tab")
    ws = wb["vancianRef"]
    grid = [[c.value for c in row] for row in ws.iter_rows()]
    dropdown = boundaries(wb, DROPDOWN_NAME)
    lookup = boundaries(wb, LOOKUP_NAME)
    wb.close()

    def at(r, c):
        """One cell, by 1-based sheet coordinates."""
        if 1 <= r <= len(grid) and 1 <= c <= len(grid[r - 1]):
            return grid[r - 1][c - 1]
        return None

    width = max((len(row) for row in grid), default=0)

    # A block starts wherever a spell-level row says "0th"; the label rows are
    # the ones where that happens, and each names a table.
    label_rows = sorted({
        r for r in range(1, len(grid) + 1)
        for c in range(1, width + 1)
        if text(at(r, c)) == BLOCK_MARKER
    })
    if len(label_rows) != len(TABLES):
        sys.exit(f"expected {len(TABLES)} stacked tables, found {len(label_rows)}: {label_rows}")

    # Columns come from the first table's label row and are shared by all three
    # -- the lower tables repeat the headers but the sheet reads them by
    # offsetting down the same column.
    starts = [c for c in range(1, width + 1)
              if text(at(label_rows[0], c)) == BLOCK_MARKER]

    classes = []
    columns = {}
    for col in starts:
        name = COLUMN_SUFFIX.sub("", text(at(label_rows[0] - 1, col)))
        if not name:
            continue
        entry = {"name": name}
        for table, label_row in zip(TABLES, label_rows):
            rows = [[number(at(label_row + 1 + lvl, col + s))
                     for s in range(SPELL_LEVELS)]
                    for lvl in range(CLASS_LEVELS)]
            if any(v is not None for row in rows for v in row):
                entry[table] = rows
        classes.append(entry)
        columns[name.lower()] = col

    return classes, reachability(grid, classes, columns, dropdown, lookup)


def reachability(grid, classes, columns, dropdown, lookup):
    """Which blocks the sheet's own dropdown and lookup range could still see.

    Classes were appended to the tab without extending either named range, so
    the tail of the table went dead on the sheet: a block past the end of
    `VancianLookup` fails its HLOOKUP and the block silently reads zero. All of
    them are extracted here regardless, so this is only worth reporting.
    """
    offered = set()
    if dropdown:
        min_col, min_row, max_col, max_row = dropdown
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                if 1 <= r <= len(grid) and 1 <= c <= len(grid[r - 1]):
                    offered.add(text(grid[r - 1][c - 1]).lower())

    lookup_max_col = lookup[2] if lookup else None
    out = {"undropped": [], "unlookupable": []}
    for c in classes:
        key = c["name"].lower()
        if dropdown and key not in offered:
            out["undropped"].append(c["name"])
        if lookup_max_col is not None and columns[key] > lookup_max_col:
            out["unlookupable"].append(c["name"])
    return out


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", help="any workbook built on the campaign template")
    parser.add_argument("-o", "--out", default="data/vancian.json")
    args = parser.parse_args()

    classes, reach = read_tables(args.workbook)
    os.makedirs(os.path.dirname(args.out) or ".", exist_ok=True)
    with open(args.out, "w", encoding="utf-8") as f:
        json.dump({"spellLevels": SPELL_LEVEL_LABELS, "classes": classes},
                  f, separators=(",", ":"), ensure_ascii=False)

    size = os.path.getsize(args.out)
    print(f"{len(classes)} classes -> {args.out} ({size:,} b)")
    with_known = sum(1 for c in classes if "known" in c)
    with_bonus = [c["name"] for c in classes if "bonus" in c]
    print(f"  {with_known} carry a spells-known table")
    print(f"  bonus slots: {', '.join(with_bonus) or 'none'}")
    if reach["undropped"]:
        print(f"  in the tab but not in its dropdown: {', '.join(reach['undropped'])}")
    if reach["unlookupable"]:
        print(f"  past the end of {LOOKUP_NAME}, so dead on the sheet: "
              f"{', '.join(reach['unlookupable'])}")


if __name__ == "__main__":
    main()
